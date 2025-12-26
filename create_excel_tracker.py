from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "MODI Project Tracker"

# Colors
dark_bg = PatternFill(start_color="1E1E2E", end_color="1E1E2E", fill_type="solid")
header_bg = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
title_bg = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
complete_bg = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
pending_bg = PatternFill(start_color="7F2315", end_color="7F2315", fill_type="solid")
summary_bg = PatternFill(start_color="CBA6F7", end_color="CBA6F7", fill_type="solid")
critical_bg = PatternFill(start_color="9D0208", end_color="9D0208", fill_type="solid")
high_bg = PatternFill(start_color="DC2F02", end_color="DC2F02", fill_type="solid")
medium_bg = PatternFill(start_color="E85D04", end_color="E85D04", fill_type="solid")
low_bg = PatternFill(start_color="F48C06", end_color="F48C06", fill_type="solid")

# Fonts
title_font = Font(name='Segoe UI', size=18, bold=True, color="00D4FF")
header_font = Font(name='Segoe UI', size=11, bold=True, color="FFFFFF")
data_font = Font(name='Segoe UI', size=10, color="CDD6F4")
module_font = Font(name='Segoe UI', size=10, bold=True, color="89B4FA")
date_font = Font(name='Segoe UI', size=10, color="FAB387")
hours_font = Font(name='Segoe UI', size=10, bold=True, color="A6E3A1")
complete_font = Font(name='Segoe UI', size=10, bold=True, color="40916C")
pending_font = Font(name='Segoe UI', size=10, bold=True, color="FCA311")
summary_font = Font(name='Segoe UI', size=11, bold=True, color="1E1E2E")

# Border
thin_border = Border(
    left=Side(style='thin', color='45475A'),
    right=Side(style='thin', color='45475A'),
    top=Side(style='thin', color='45475A'),
    bottom=Side(style='thin', color='45475A')
)

center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')

# Column widths
col_widths = [6, 35, 30, 14, 14, 10, 15, 12, 25]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Title Row
ws.merge_cells('A1:I1')
ws['A1'] = "📊 MODI PROJECT - WORKING TRACKER"
ws['A1'].font = title_font
ws['A1'].fill = title_bg
ws['A1'].alignment = center_align

# Subtitle Row
ws.merge_cells('A2:I2')
ws['A2'] = "Medical OPD Digital Interface | Nov 2025 - Dec 2025 | Updated: 15-Dec-2025"
ws['A2'].font = Font(name='Segoe UI', size=10, color="E94560")
ws['A2'].fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
ws['A2'].alignment = center_align

# Headers
headers = ["Sr#", "📁 Module/File", "📝 Description", "📅 Start", "📅 End", "⏱️ Hrs", "📊 Status", "⚡ Priority", "📌 Notes"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_bg
    cell.alignment = center_align
    cell.border = thin_border

# Data
data = [
    [1, "main.dart", "App Entry Point & Routing", "01-Nov-2025", "02-Nov-2025", 4, "✅ COMPLETE", "HIGH", "Theme & navigation setup"],
    [2, "splash_screen.dart", "Animated Splash Screen", "02-Nov-2025", "03-Nov-2025", 6, "✅ COMPLETE", "MEDIUM", "Custom logo animations"],
    [3, "login_signup_choice.dart", "Login/Signup Selection", "03-Nov-2025", "05-Nov-2025", 8, "✅ COMPLETE", "HIGH", "Role-based options"],
    [4, "doctor_login_page.dart", "Doctor Login Auth", "05-Nov-2025", "07-Nov-2025", 10, "✅ COMPLETE", "HIGH", "Secure validation"],
    [5, "staff_login_page.dart", "Staff Login Page", "07-Nov-2025", "08-Nov-2025", 6, "✅ COMPLETE", "HIGH", "OPD staff auth"],
    [6, "doctor_registration_page.dart", "Doctor Registration", "08-Nov-2025", "12-Nov-2025", 16, "✅ COMPLETE", "HIGH", "Multi-step form"],
    [7, "doctor_dashboard.dart", "Main Doctor Dashboard", "12-Nov-2025", "25-Nov-2025", 40, "✅ COMPLETE", "CRITICAL", "Full analytics dashboard"],
    [8, "opd_staff_dashboard.dart", "Staff Dashboard", "20-Nov-2025", "30-Nov-2025", 35, "✅ COMPLETE", "CRITICAL", "Patient queue mgmt"],
    [9, "patient_registration_form.dart", "Patient Registration", "25-Nov-2025", "02-Dec-2025", 30, "✅ COMPLETE", "CRITICAL", "Photo & validation"],
    [10, "patient_detail_view.dart", "Patient Details View", "01-Dec-2025", "08-Dec-2025", 32, "✅ COMPLETE", "HIGH", "Medical history view"],
    [11, "consultation_screen.dart", "Doctor Consultation", "02-Dec-2025", "06-Dec-2025", 18, "✅ COMPLETE", "HIGH", "Diagnosis & Rx"],
    [12, "prescription_page.dart", "Prescription Generation", "04-Dec-2025", "07-Dec-2025", 12, "✅ COMPLETE", "HIGH", "PDF prescription"],
    [13, "payment_management.dart", "Payment System", "05-Dec-2025", "08-Dec-2025", 20, "✅ COMPLETE", "HIGH", "Multi-mode payments"],
    [14, "payment_installment_screen.dart", "Installment Management", "06-Dec-2025", "10-Dec-2025", 24, "✅ COMPLETE", "MEDIUM", "EMI tracking"],
    [15, "database_helper.dart", "SQLite Database Layer", "01-Nov-2025", "10-Dec-2025", 50, "✅ COMPLETE", "CRITICAL", "All CRUD operations"],
    [16, "models.dart", "Data Models", "01-Nov-2025", "10-Nov-2025", 15, "✅ COMPLETE", "HIGH", "Patient/Doctor/Staff"],
    [17, "patient_qr_code.dart", "QR Code Generation", "08-Dec-2025", "12-Dec-2025", 18, "✅ COMPLETE", "MEDIUM", "Deep link QR"],
    [18, "waiting_room_display.dart", "TV Display Queue", "10-Dec-2025", "14-Dec-2025", 22, "✅ COMPLETE", "MEDIUM", "Large screen"],
    [19, "whatsapp_integration.dart", "WhatsApp Messaging", "08-Dec-2025", "12-Dec-2025", 16, "✅ COMPLETE", "MEDIUM", "Patient notifications"],
    [20, "sms_integration.dart", "SMS Service", "09-Dec-2025", "12-Dec-2025", 14, "✅ COMPLETE", "MEDIUM", "Bulk SMS"],
    [21, "birthday_notification_widget.dart", "Birthday Reminders", "08-Dec-2025", "10-Dec-2025", 10, "✅ COMPLETE", "LOW", "Auto birthday wishes"],
    [22, "reports_analytics.dart", "Analytics Dashboard", "06-Dec-2025", "10-Dec-2025", 18, "✅ COMPLETE", "MEDIUM", "Revenue charts"],
    [23, "settings_configuration.dart", "App Settings", "05-Dec-2025", "10-Dec-2025", 16, "✅ COMPLETE", "MEDIUM", "Clinic config"],
    [24, "pdf_service.dart", "PDF Generation", "04-Dec-2025", "08-Dec-2025", 20, "✅ COMPLETE", "HIGH", "Invoices & Rx"],
    [25, "design_system.dart", "UI Design System", "01-Nov-2025", "15-Nov-2025", 12, "✅ COMPLETE", "HIGH", "Colors/fonts/themes"],
    [26, "responsive_helper.dart", "Responsive Layouts", "10-Nov-2025", "15-Nov-2025", 10, "✅ COMPLETE", "MEDIUM", "Mobile/Desktop"],
    [27, "glassmorphism.dart", "Glass UI Effects", "12-Nov-2025", "14-Nov-2025", 6, "✅ COMPLETE", "LOW", "Modern UI styling"],
    [28, "appointment_management.dart", "Appointment System", "01-Dec-2025", "06-Dec-2025", 14, "✅ COMPLETE", "HIGH", "Booking & scheduling"],
    [29, "book_appointment.dart", "Appointment Booking", "02-Dec-2025", "05-Dec-2025", 12, "✅ COMPLETE", "HIGH", "Patient booking UI"],
    [30, "follow_up_appointments.dart", "Follow-up Tracking", "05-Dec-2025", "08-Dec-2025", 10, "✅ COMPLETE", "MEDIUM", "Auto reminders"],
    [31, "patient_search.dart", "Patient Search", "28-Nov-2025", "02-Dec-2025", 12, "✅ COMPLETE", "HIGH", "Name/Mobile/QR"],
    [32, "notifications_center.dart", "Notification Hub", "06-Dec-2025", "10-Dec-2025", 10, "✅ COMPLETE", "MEDIUM", "Central alerts"],
    [33, "medicine_database.dart", "Medicine Catalog", "01-Dec-2025", "06-Dec-2025", 18, "✅ COMPLETE", "MEDIUM", "Drug database"],
    [34, "staff_management.dart", "Staff CRUD", "25-Nov-2025", "30-Nov-2025", 12, "✅ COMPLETE", "MEDIUM", "Staff accounts"],
    [35, "indian_festival_calendar.dart", "Festival Calendar", "08-Dec-2025", "10-Dec-2025", 6, "✅ COMPLETE", "LOW", "Holiday planning"],
    [36, "patient_feedback_system.dart", "Feedback QR", "14-Dec-2025", "14-Dec-2025", 4, "✅ COMPLETE", "LOW", "Google Form QR"],
    [37, "forgot_password_page.dart", "Password Reset", "08-Nov-2025", "10-Nov-2025", 8, "✅ COMPLETE", "HIGH", "Email reset flow"],
    [38, "advertisement_page.dart", "Clinic Ads Display", "10-Dec-2025", "12-Dec-2025", 10, "✅ COMPLETE", "LOW", "Promotional content"],
    [39, "patient_report_page.dart", "Patient Report View", "12-Dec-2025", "14-Dec-2025", 8, "✅ COMPLETE", "MEDIUM", "Web-based report"],
    [40, "voice_prescription.dart", "Voice Input", "04-Dec-2025", "-", 6, "⏳ PENDING", "LOW", "Speech to text"],
    [41, "lab_reports_management.dart", "Lab Reports", "10-Dec-2025", "-", 8, "⏳ PENDING", "MEDIUM", "Report uploads"],
    [42, "email_service.dart", "Email Integration", "08-Dec-2025", "10-Dec-2025", 6, "✅ COMPLETE", "MEDIUM", "SMTP service"],
    [43, "🔮 Multi-clinic Support", "Multiple Branches", "-", "-", "-", "⏳ PENDING", "HIGH", "Future enhancement"],
    [44, "🔮 Inventory Management", "Medicine Stock", "-", "-", "-", "⏳ PENDING", "MEDIUM", "Future enhancement"],
    [45, "🔮 Insurance Integration", "Claim Processing", "-", "-", "-", "⏳ PENDING", "LOW", "Future enhancement"],
]

for row_idx, row_data in enumerate(data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = dark_bg
        cell.border = thin_border
        cell.alignment = center_align if col_idx != 2 and col_idx != 9 else left_align
        
        if col_idx == 1:
            cell.font = data_font
        elif col_idx == 2:
            cell.font = module_font
        elif col_idx in [4, 5]:
            cell.font = date_font
        elif col_idx == 6:
            cell.font = hours_font
        elif col_idx == 7:
            if "COMPLETE" in str(value):
                cell.font = complete_font
                cell.fill = complete_bg
            else:
                cell.font = pending_font
                cell.fill = pending_bg
        elif col_idx == 8:
            cell.font = Font(name='Segoe UI', size=10, bold=True, color="FFFFFF")
            if value == "CRITICAL":
                cell.fill = critical_bg
            elif value == "HIGH":
                cell.fill = high_bg
            elif value == "MEDIUM":
                cell.fill = medium_bg
            elif value == "LOW":
                cell.fill = low_bg
        else:
            cell.font = data_font

# Summary Row
summary_row = len(data) + 4
ws.merge_cells(f'A{summary_row}:E{summary_row}')
ws[f'A{summary_row}'] = "📈 PROJECT SUMMARY"
ws[f'A{summary_row}'].font = summary_font
ws[f'A{summary_row}'].fill = summary_bg
ws[f'A{summary_row}'].alignment = center_align
ws[f'A{summary_row}'].border = thin_border

ws[f'F{summary_row}'] = "609 HRS"
ws[f'F{summary_row}'].font = summary_font
ws[f'F{summary_row}'].fill = summary_bg
ws[f'F{summary_row}'].alignment = center_align
ws[f'F{summary_row}'].border = thin_border

ws[f'G{summary_row}'] = "40✅ | 5⏳"
ws[f'G{summary_row}'].font = summary_font
ws[f'G{summary_row}'].fill = summary_bg
ws[f'G{summary_row}'].alignment = center_align
ws[f'G{summary_row}'].border = thin_border

ws.merge_cells(f'H{summary_row}:I{summary_row}')
ws[f'H{summary_row}'] = "88.9% COMPLETE"
ws[f'H{summary_row}'].font = summary_font
ws[f'H{summary_row}'].fill = summary_bg
ws[f'H{summary_row}'].alignment = center_align
ws[f'H{summary_row}'].border = thin_border

# Row heights
ws.row_dimensions[1].height = 35
ws.row_dimensions[2].height = 25
ws.row_dimensions[3].height = 30
for i in range(4, summary_row + 1):
    ws.row_dimensions[i].height = 22

wb.save("e:/modi/MODI_PROJECT_TRACKER.xlsx")
print("✅ Excel file created: MODI_PROJECT_TRACKER.xlsx")
